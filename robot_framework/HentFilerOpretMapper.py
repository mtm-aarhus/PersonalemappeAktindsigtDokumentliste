from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from OpenOrchestrator.database.queues import QueueElement
import os
import pandas as pd
import re
import xml.etree.ElementTree as ET
import requests
import json
from urllib.parse import quote
from datetime import datetime, timedelta
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.sharing.links.kind import SharingLinkKind
from office365.sharepoint.webs.web import Web
from requests_ntlm import HttpNtlmAuth
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, Protection
from PIL import ImageFont, ImageDraw, Image
from urllib.parse import urlencode

def HentFilerOpretMapper(caseid, PersonaleSagsID: str, SagsID: str, MappeNavn, GOAPI_URL, GOAPILIVECRED_username, GOAPILIVECRED_password, RobotUsername, RobotPassword, SharepointURL, orchestrator_connection):
    '''
    Folder that takes the case ID of a case and creates folders and subfolders in sharepoint containing document lists
    '''
    url = GOAPI_URL + "/_goapi/Cases/Metadata/" + SagsID
    
    session = requests.Session()
    session.auth = HttpNtlmAuth(GOAPILIVECRED_username, GOAPILIVECRED_password)
    session.headers.update({"Content-Type": "application/json"})
    response = session.get(url, timeout=500)
    response.raise_for_status()

    SagMetaData = response.text
    json_obj = json.loads(SagMetaData)

    metadata_xml = json_obj.get("Metadata")

    if metadata_xml:
        xdoc = ET.fromstring(metadata_xml)
        SagsURL = xdoc.attrib.get("ows_CaseUrl")
        SagsTitel = xdoc.attrib.get("ows_Title")

        if SagsURL and "cases/" in SagsURL:
            Akt = SagsURL.split("cases/")[1].split("/")[0]
        else:
            print("Error: 'cases/' not found in SagsURL or SagsURL is missing.")
    else:
        print("Error: 'Metadata' field is missing in the JSON response.")

    SagsTitel = re.sub(r'[~#%&*{}\:\\<>?/+|\"\'\t\[\]`^@=!$();\€£¥₹]', '', str(SagsTitel))
    SagsTitel = " ".join(SagsTitel.split())

    columns = [
        "Akt ID", "Dok ID", "Dokumenttitel", "Dokumentkategori", "Dokumentdato", 
        "Bilag til Dok ID", "Bilag", "Link til dokument", 
        "Omfattet af ansøgningen? (Ja/Nej)", "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)", 
        "Begrundelse hvis nej eller delvis"
    ]
    data_table = pd.DataFrame(columns=columns)

    Akt = SagsURL.split("/")[1]  
    encoded_sags_id = SagsID.rsplit('-', 1)[0].replace("-", "%2D")
    endelse =  SagsID.rsplit('-', 1)[-1]
    ListURL = f"%27%2Fcases%2F{Akt}%2F{encoded_sags_id}%2FDokumenter%27"

    response = session.get(f"{GOAPI_URL}/{SagsURL}/_goapi/Administration/GetLeftMenuCounter/{endelse}")

    ViewsIDArray = json.loads(response.text) # Parse the JSON

    for item in ViewsIDArray:
        if item["ViewName"] == "IkkeJournaliseret.aspx":
            print('Den er ikke journaliseret ')
            ikke_journaliseret_id = item["ViewId"]  
        elif item["ViewName"] == "Journaliseret.aspx":
            print('Den er journaliseret ')
            journaliseret_id = item["ViewId"]

    view_ids_to_use = [ikke_journaliseret_id, journaliseret_id]

    for current_view_id in view_ids_to_use:
        firstrun = True
        MorePages = True

        while MorePages:
            print("Henter dokumentlister")

            # If not the first run, fetch the next page
            if not firstrun:
                url = f"{GOAPI_URL}/{SagsURL}/_api/web/GetList(@listUrl)/RenderListDataAsStream"
                url_with_query = f"{url}?@listUrl={ListURL}{NextHref.replace('?', '&')}"

                response = session.post(url_with_query, timeout=500)
                response.raise_for_status()
                Dokumentliste = response.text  # Extract the content
            else:
                # If first run, fetch the first page for the current view
                url = f"{GOAPI_URL}/{SagsURL}/_api/web/GetList(@listUrl)/RenderListDataAsStream"
                query_params = f"?@listUrl={ListURL}&View={current_view_id}&RootFolder=%2Fcases%2F{Akt}%2F{encoded_sags_id}%2FDokumenter%2F{endelse}&FilterField1=CCMSubID&FilterValue1={endelse}"
                full_url = url + query_params

                response = session.post(full_url, timeout=500)
                response.raise_for_status()
                Dokumentliste = response.text  # Extract the content

            # Deserialize response
            dokumentliste_json = json.loads(Dokumentliste)
            dokumentliste_rows = dokumentliste_json.get("Row", [])

            # Check for additional pages
            NextHref = dokumentliste_json.get("NextHref")
            MorePages = "NextHref" in dokumentliste_json

            # Process each row
            for item in dokumentliste_rows:
                # Extract and prepare data
                DokumentURL = GOAPI_URL + quote(item.get("FileRef", ""), safe="/")
                AktID = item.get("CaseRecordNumber", "").replace(".", "")
                DokumentDato = str(item.get("Dato"))
                Dokumenttitel = item.get("Title", "")
                DokID = str(item.get("DocID"))
                DokumentKategori = str(item.get("Korrespondance"))

                if len(Dokumenttitel) < 2:
                    Dokumenttitel = item.get("FileLeafRef.Name", "")

                # Fetch parents and children data
                parents_response = session.get(f"{GOAPI_URL}/_goapi/Documents/Parents/{DokID}", timeout=500)
                parents_object = json.loads(parents_response.text)
                ParentArray = parents_object.get("ParentsData", [])
                Bilag = ", ".join(str(currentItem.get("DocumentId", "")) for currentItem in ParentArray)

                children_response = session.get(f"{GOAPI_URL}/_goapi/Documents/Children/{DokID}", timeout=500)
                children_object = json.loads(children_response.text)
                ChildrenArray = children_object.get("ChildrenData", [])
                BilagChild = ", ".join(str(currentItem.get("DocumentId", "")) for currentItem in ChildrenArray)

                # Append data to DataFrame
                if "tunnel_marking" in Dokumenttitel.lower() or "memometadata" in Dokumenttitel.lower():
                    memo_tunnel = True
                    data_table = pd.concat([data_table, pd.DataFrame([{
                        "Akt ID": AktID,
                        "Dok ID": DokID,
                        "Dokumenttitel": Dokumenttitel,
                        "Dokumentkategori": DokumentKategori,
                        "Dokumentdato": DokumentDato,
                        "Bilag": BilagChild,
                        "Bilag til Dok ID": Bilag,
                        "Link til dokument": DokumentURL,
                        "Omfattet af ansøgningen? (Ja/Nej)": "Ja",
                        "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)": "Nej",
                        "Begrundelse hvis nej eller delvis": "Tavshedsbelagte oplysninger - om private forhold"
                    }])], ignore_index=True)
                else:
                    data_table = pd.concat([data_table, pd.DataFrame([{
                        "Akt ID": AktID,
                        "Dok ID": DokID,
                        "Dokumenttitel": Dokumenttitel,
                        "Dokumentkategori": DokumentKategori,
                        "Dokumentdato": DokumentDato,
                        "Bilag": BilagChild,
                        "Bilag til Dok ID": Bilag,
                        "Link til dokument": DokumentURL,
                        "Omfattet af ansøgningen? (Ja/Nej)": "Ja",
                        "Gives der aktindsigt i dokumentet? (Ja/Nej/Delvis)": "",
                        "Begrundelse hvis nej eller delvis": ""
                    }])], ignore_index=True)

            firstrun = False


    # Define font settings
    FONT_PATH = "calibri.ttf"  # Ensure this file exists in your directory
    FONT_SIZE = 11

    # Load the font
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
    except OSError:
        raise FileNotFoundError(f"Font file not found at {FONT_PATH}. Please ensure the font file is available.")

    # Function to calculate text dimensions in Excel units
    def calculate_text_dimensions(text, font, max_width_in_pixels):
        dummy_image = Image.new("RGB", (1, 1))
        draw = ImageDraw.Draw(dummy_image)
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        excel_column_width = text_width / 5
        lines = max(1, text_width // max_width_in_pixels + 1)
        excel_row_height = lines * (text_height / 1.33)
        return excel_column_width, excel_row_height

    if data_table.empty:
        fake_row = {col: "" for col in data_table.columns}
        data_table = pd.DataFrame([fake_row])  # Add placeholder row

    # Ensure 'Akt ID' is numeric and clean
    data_table['Akt ID'] = pd.to_numeric(data_table['Akt ID'].astype(str).str.strip(), errors='coerce')
    print('Doing data table stuff')

    # Sort values if the table is not empty
    if not data_table.empty:
        data_table = data_table.sort_values(by='Akt ID', ascending=True, ignore_index=True)

    # 🟢 STEP 2: SAVE THE DATAFRAME TO EXCEL
    excel_file_path = f"{SagsID}_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    data_table.to_excel(excel_file_path, index=False, sheet_name="Sagsoversigt")

    # Open Excel file for formatting
    workbook = load_workbook(excel_file_path)
    worksheet = workbook["Sagsoversigt"]

    # Ensure at least 2 rows exist (header + data row)
    if worksheet.max_row == 1:
        worksheet.append([""] * worksheet.max_column)  # Add an empty row

    data_range = f"A1:K{worksheet.max_row}"
    table = Table(displayName="SagsoversigtTable", ref=data_range)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Apply column width formatting dynamically
    max_width_in_pixels = 382
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                text = str(cell.value)
                column_width, _ = calculate_text_dimensions(text, font, max_width_in_pixels)
                max_length = max(max_length, column_width)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    # Specific column adjustments
    COLUMN_C_INDEX, COLUMN_G_INDEX = 3, 7
    worksheet.column_dimensions[get_column_letter(COLUMN_C_INDEX)].width = 50

    # Define header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = header_font

    # Apply row height adjustments for wrapped text
    ROW_HEIGHT_PER_PIXEL = 1
    def calculate_row_height(text, font, max_width_in_pixels):
        dummy_image = Image.new("RGB", (1, 1))
        draw = ImageDraw.Draw(dummy_image)
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        lines_required = max(1, (text_width / max_width_in_pixels) + 1)
        return lines_required * text_height * ROW_HEIGHT_PER_PIXEL

    # Adjust row heights for Columns C and G
    for row_idx in range(2, worksheet.max_row + 1):
        row_height = 15
        for col_idx in [COLUMN_C_INDEX, COLUMN_G_INDEX]:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)
                text = str(cell.value)
                height = calculate_row_height(text, font, 150 if col_idx == COLUMN_C_INDEX else 70)
                row_height = max(row_height, height)
        worksheet.row_dimensions[row_idx].height = row_height

    for col in ["I", "J", "K"]:
        for row_idx in range(2, worksheet.max_row + 1):  # Skip header
            cell = worksheet[f"{col}{row_idx}"]
            cell.protection = Protection(locked=False)  # Allow dropdown selection

    # Add hyperlinks in column H
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=8)
        if cell.value:
            cell.value, cell.hyperlink, cell.style = "Dokumentlink", cell.value, "Hyperlink"

    # Add dropdown validations
    validation_i = DataValidation(type="list", formula1='"Ja,Nej"', allow_blank=False, showErrorMessage=True)
    validation_i.error, validation_i.errorTitle = "Vælg venligst Ja eller Nej.", "Ugyldig værdi"

    validation_j = DataValidation(type="list", formula1='"Ja,Delvis,Nej"', allow_blank=False, showErrorMessage=True)
    validation_j.error, validation_j.errorTitle = "Vælg venligst Ja, Delvis eller Nej.", "Ugyldig værdi"

    # Create hidden sheet for dropdown options
    hidden_options = [
        "Internt dokument - ufærdigt arbejdsdokument",
        "Internt dokument - foreløbige og sagsforberedende overvejelser",
        "Internt dokument - del af intern beslutningsproces",
        "Særlige dokumenter - korrespondance med sagkyndig rådgiver vedr. tvistsag",
        "Særlige dokumenter - statistik og undersøgelser",
        "Særlige dokumenter - straffesag",
        "Tavshedsbelagte oplysninger - om private forhold",
        "Tavshedsbelagte oplysninger - forretningsforhold",
        "Tavshedsbelagte oplysninger - Andet (uddybes i afgørelsen)",
        " "
    ]

    hidden_sheet = workbook.create_sheet("VeryHidden")
    hidden_sheet.sheet_state = "veryHidden"
    for idx, option in enumerate(hidden_options, start=1):
        hidden_sheet.cell(row=idx, column=1, value=option)

    # Add validation for column K using hidden sheet values
    validation_k = DataValidation(type="list", formula1=f"=VeryHidden!$A$1:$A${len(hidden_options)}",
                                allow_blank=False, showErrorMessage=True)
    validation_k.error, validation_k.errorTitle = "Vælg en mulighed.", "Ugyldig indtastning"

    first_data_row = 2 if worksheet.max_row > 1 else 1
    validation_i.add(f"I{first_data_row}:I{worksheet.max_row}")
    validation_j.add(f"J{first_data_row}:J{worksheet.max_row}")
    validation_k.add(f"K{first_data_row}:K{worksheet.max_row}")

    worksheet.add_data_validation(validation_i)
    worksheet.add_data_validation(validation_j)
    worksheet.add_data_validation(validation_k)

    worksheet.protection.sheet = True
    worksheet.protection.password = "Aktbob"
    worksheet.protection.enable()

    workbook.save(excel_file_path)

    Mappe1 = str(caseid) + " - " + str(PersonaleSagsID) + " - Personaleaktindsigtsanmodning"
    Mappe2 = str(SagsID) + '-' + SagsTitel

    # Authenticate to SharePoint using Office365 credentials
    certification = orchestrator_connection.get_credential("SharePointCert")
    api = orchestrator_connection.get_credential("SharePointAPI")

    cert_credentials = {
        "tenant": api.username,
        "client_id": api.password,
        "thumbprint": certification.username,
        "cert_path": certification.password
    }

    ctx = ClientContext(SharepointURL).with_client_certificate(**cert_credentials)

    # Function to sanitize folder names
    def sanitize_folder_name(folder_name):
        pattern = r'[.,~#%&*{}\[\]\\:<>?/+|$¤£€\"\t]'
        folder_name = re.sub(pattern, "", folder_name)
        folder_name = re.sub(r"\s+", " ", folder_name).strip()
        return folder_name

    # Sanitize folder names
    Mappe1 = sanitize_folder_name(Mappe1)
    Mappe2 = sanitize_folder_name(Mappe2)

    # Ensure folder names don't exceed length limits
    if len(Mappe1) > 99:
        Mappe1 = Mappe1[:95] + "(...)"
    if len(Mappe2) > 99:
        Mappe2 = Mappe2[:95] + "(...)"

    total_length = len(Mappe1) + len(Mappe2) + 17  # 17 is for folder structure overhead
    if total_length > 400:
        excess_length = total_length - 400
        half_excess = excess_length // 2
        Mappe1 = Mappe1[:len(Mappe1) - half_excess - 5] + "(...)"
        Mappe2 = Mappe2[:len(Mappe2) - half_excess - 5] + "(...)"

    parent_folder_name = SharepointURL.split(".com")[-1] + "/Delte dokumenter/Dokumentlister"

        # Create main folder
    root_folder = ctx.web.get_folder_by_server_relative_url(parent_folder_name)
    main_folder = root_folder.folders.add(Mappe1) 
    ctx.execute_query()

    # Create subfolder inside main folder
    subfolder = main_folder.folders.add(Mappe2)
    ctx.execute_query()

    file_path = excel_file_path  # Ensure it points to the created Excel file

    # Check if the file exists and upload it
    if os.path.exists(file_path):
        with open(file_path, "rb") as file_content:
            subfolder.upload_file(os.path.basename(file_path), file_content.read())
        ctx.execute_query()
        os.remove(file_path)
    else:
        print(f"File '{file_path}' does not exist.")
